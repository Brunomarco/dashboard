# GMA Innovation Lab - Executive Financial Intelligence System
# Full MBB-Level Dashboard | 9 Analysis Views | 70+ Metrics
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

st.set_page_config(page_title="GMA Innovation Lab - Financial Intelligence",page_icon="\u25c8",layout="wide",initial_sidebar_state="expanded")

NAVY="#1B2A4A";ACCENT="#4A6FA5";SLATE="#6B7280";LIGHT="#F8F9FB";BORDER="#E5E7EB"
GREEN="#059669";RED="#DC2626";AMBER="#D97706";TEAL="#0D9488";PURPLE="#7C3AED"

st.markdown('''
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="st-"]{font-family:"Inter",-apple-system,sans-serif}
.main .block-container{padding:2rem 3rem;max-width:1440px}
h1,h2,h3{color:#1B2A4A !important;font-weight:700 !important}
h1{font-size:26px !important;letter-spacing:-0.5px !important}
h2{font-size:17px !important;margin-top:1.5rem !important}
h3{font-size:14px !important}
[data-testid="stMetric"]{background:#F8F9FB;border:1px solid #E5E7EB;border-radius:8px;padding:18px 22px}
[data-testid="stMetricLabel"]{font-size:10px !important;font-weight:600 !important;letter-spacing:1.5px !important;text-transform:uppercase !important;color:#6B7280 !important}
[data-testid="stMetricValue"]{font-size:24px !important;font-weight:700 !important;color:#1B2A4A !important}
[data-testid="stMetricDelta"]>div{font-size:11px !important}
[data-testid="stSidebar"]{background:#1B2A4A}
[data-testid="stSidebar"] *{color:rgba(255,255,255,0.85) !important}
[data-testid="stSidebar"] .stSelectbox label,[data-testid="stSidebar"] .stRadio label{font-size:10px !important;font-weight:600 !important;letter-spacing:1.5px !important;text-transform:uppercase !important}
[data-testid="stSidebar"] hr{border-color:rgba(255,255,255,0.1) !important}
.tag{font-size:10px;font-weight:600;letter-spacing:2.5px;text-transform:uppercase;color:#4A6FA5}
.divider{border:none;border-top:1px solid #E5E7EB;margin:1.2rem 0}
.insight{background:#F8F9FB;border-left:3px solid #4A6FA5;padding:14px 18px;border-radius:0 6px 6px 0;margin:10px 0;font-size:13px;color:#3A3F47;line-height:1.55}
#MainMenu,footer,header{visibility:hidden}
</style>
''',unsafe_allow_html=True)

@st.cache_data
def load_data():
    np.random.seed(42)
    months=pd.date_range('2024-01-01',periods=12,freq='MS')
    rev=np.array([420,445,460,485,470,510,530,520,555,540,570,590])*1000
    rev_prod=np.array([252,267,276,291,282,306,318,312,333,324,342,354])*1000
    rev_svc=np.array([126,133.5,138,145.5,141,153,159,156,166.5,162,171,177])*1000
    rev_oth=rev-rev_prod-rev_svc
    cg=np.array([243600,253650,271400,271600,272600,290700,296800,301600,305250,307800,319200,324500])
    cg_mat=np.array([170520,177555,189980,190120,190820,203490,207760,211120,213675,215460,223440,227150])
    cg_lab=np.array([48720,50730,54280,54320,54520,58140,59360,60320,61050,61560,63840,64900])
    cg_oh=cg-cg_mat-cg_lab
    gp=rev-cg
    pers=np.array([52,52,53,53,53,54,54,54,55,55,55,56])*1000
    rnt=np.full(12,12000);mkt=np.array([8,9,7.5,10,8.5,9.5,11,9,10.5,9,10,11.5])*1000
    trav=np.array([3,2.5,3.5,4,2,3.5,4.5,3,4,2.5,3,4])*1000
    tech=np.array([5,5,5.2,5.2,5.2,5.4,5.4,5.4,5.6,5.6,5.6,5.8])*1000
    prof=np.array([4,3.5,4.5,3,3.5,4,2.5,4.5,3,3.5,4,3])*1000
    oox=np.array([3,3.5,2.8,3.3,2.8,3.6,2.6,3.6,2.9,2.9,2.9,3.2])*1000
    topx=pers+rnt+mkt+trav+tech+prof+oox
    ebitda=gp-topx;dep=np.full(12,5000);amort=np.full(12,3000);da=dep+amort
    ebit=ebitda-da;intr=np.linspace(3500,2400,12).astype(int)
    oinc=np.array([500,800,300,600,400,700,500,400,800,300,600,500])
    ebt=ebit-intr+oinc;tx=np.maximum(0,(ebt*0.25)).astype(int);ni=ebt-tx
    # Balance Sheet
    cash=np.array([180,195,188,210,205,225,240,235,255,248,265,285])*1000
    sinv=np.full(12,20000)
    rec_t=np.array([105,112,118,125,120,132,138,135,146,141,150,156])*1000
    rec_o=np.full(12,20000);rec=rec_t+rec_o
    inv_r=np.array([38,39.2,40.8,40,42,43.2,44,42.4,44.8,46,47.2,48])*1000
    inv_w=inv_r/2;inv_f=inv_r.copy();inv=inv_r+inv_w+inv_f
    pre=np.full(12,8000);oca=np.full(12,5000)
    tca=cash+sinv+rec+inv+pre+oca
    pay_t=np.array([72,75,78,77,81,84,86,82,87,90,92,94])*1000
    pay_o=np.array([13,13,14,13,14,14,14,14,15,15,16,16])*1000
    pay=pay_t+pay_o;acc=np.linspace(18000,22000,12).astype(int)
    sd=np.linspace(50000,28000,12).astype(int);ctx=np.full(12,4000);ocl=np.full(12,6000)
    tcl=pay+acc+sd+ctx+ocl
    ppe_n=np.linspace(320000,275000,12).astype(int);itg=np.linspace(45000,39500,12).astype(int)
    gw=np.full(12,80000);onca=np.full(12,15000);tnca=ppe_n+itg+gw+onca
    ta=tca+tnca;ld=np.linspace(200000,156000,12).astype(int)
    pens=np.full(12,15000);dtx=np.full(12,8000);oncl=np.full(12,5000)
    tncl=ld+pens+dtx+oncl;tl=tcl+tncl;eq=ta-tl
    # Cash Flow
    facs=[.82,.85,.79,.91,.83,.88,.86,.80,.93,.84,.87,.90]
    opcf=np.array([int(e*f) for e,f in zip(ebitda,facs)])
    capex=np.array([-25,-30,-18,-35,-15,-32,-25,-20,-38,-15,-28,-33])*1000
    ioth=np.array([-3,-5,-4,-6,-4,-5,-5,-5,-4,-5,-5,-5])*1000
    icf=capex+ioth
    drep=np.full(12,-4000);divs=np.array([0,0,0,0,0,0,-15000,0,0,0,0,-15000]);ofin=np.full(12,-1000)
    fcf_arr=drep+divs+ofin;ncf=opcf+icf+fcf_arr
    fcf=opcf+capex
    # Budget/PY/HC
    brev=rev*np.array([1.02,.97,1.03,.97,1.01,1.04,.99,1.02,.98,1.03,1.01,.97])
    bebitda=ebitda*np.array([1.05,.96,1.02,.98,1.04,1.01,.97,1.03,.99,1.02,1.00,.98])
    pyrev=(rev*0.88).astype(int);pyebitda=(ebitda*0.82).astype(int);pyni=(ni*0.78).astype(int)
    hc=np.array([45,45,46,46,46,47,47,47,48,48,48,49])
    # MBB Ratios
    gm=gp/rev*100;em=ebitda/rev*100;ebitm=ebit/rev*100;nm=ni/rev*100
    roe=ni/eq*100;roa=ni/ta*100
    nopat=(ebit*(1-0.25)).astype(int);ic=eq+ld+sd-cash;roic=nopat/np.where(ic>0,ic,1)*100
    dp_nm=nm/100;dp_at=rev/ta;dp_em=ta/eq;dp_roe=dp_nm*dp_at*dp_em*100
    cr=tca/tcl;qr=(tca-inv)/tcl;cashr=cash/tcl
    de=tl/eq;dta=tl/ta;nd=sd+ld-cash;nd_eb=nd/(ebitda*12);icov=ebit/np.where(intr>0,intr,1)
    dso=rec_t/rev*30;dio=inv/cg*30;dpo=pay_t/cg*30;ccc=dso+dio-dpo
    at=rev/ta;invt=cg/inv;rect=rev/rec_t
    nwc=tca-tcl;nwc_p=nwc/rev*100
    fix=pers+rnt+tech;var=cg+mkt+trav;con=rev-var;conm=con/rev*100
    be=fix/(con/rev);oplev=con/ebitda
    fcfy=fcf/rev*100;cfeb=opcf/ebitda*100;cxr=np.abs(capex)/rev*100;cxd=np.abs(capex)/dep
    altman=0.717*(nwc/ta)+0.847*((eq-100000)/ta)+3.107*(ebit/ta)+0.420*(eq/tl)+0.998*(rev/ta)
    rvar=(rev-brev)/brev*100;evar=(ebitda-bebitda)/bebitda*100
    rmom=np.concatenate([[0],np.diff(rev)/rev[:-1]*100])
    ryoy=(rev/pyrev-1)*100;emom=np.concatenate([[0],np.diff(ebitda)/ebitda[:-1]*100])
    ravg=rev.sum()/12;seas=rev/ravg
    rhc=rev/hc;ehc=ebitda/hc;ohc=topx/hc
    WACC=0.10;eva=nopat-(WACC/12)*ic
    return pd.DataFrame({
        'Month':months,'Label':[m.strftime('%b %y') for m in months],
        'Revenue':rev,'Rev_Prod':rev_prod,'Rev_Svc':rev_svc,'Rev_Oth':rev_oth,
        'COGS':cg,'COGS_Mat':cg_mat,'COGS_Lab':cg_lab,'COGS_OH':cg_oh,'GP':gp,
        'Pers':pers,'Rent':rnt,'Mkt':mkt,'Trav':trav,'Tech':tech,'Prof':prof,'OOx':oox,'TotalOpEx':topx,
        'EBITDA':ebitda,'Dep':dep,'Amort':amort,'DA':da,'EBIT':ebit,
        'Interest':intr,'OInc':oinc,'EBT':ebt,'Tax':tx,'NI':ni,
        'Cash':cash,'SInv':sinv,'Rec_T':rec_t,'Rec_O':rec_o,'Rec':rec,
        'Inv_R':inv_r,'Inv_W':inv_w,'Inv_F':inv_f,'Inv':inv,
        'TCA':tca,'PPE':ppe_n,'Intang':itg,'GW':gw,'TNCA':tnca,'TA':ta,
        'Pay_T':pay_t,'Pay':pay,'Acc':acc,'SD':sd,'TCL':tcl,
        'LD':ld,'TNCL':tncl,'TL':tl,'Eq':eq,
        'OpCF':opcf,'CapEx':capex,'InvCF':icf,'FinCF':fcf_arr,'NetCF':ncf,'FCF':fcf,
        'BRev':brev,'BEBITDA':bebitda,'PYRev':pyrev,'PYEBITDA':pyebitda,'PYNI':pyni,'HC':hc,
        'GM':gm,'EM':em,'EBITM':ebitm,'NM':nm,'ROE':roe,'ROA':roa,'ROIC':roic,
        'DuPont_NM':dp_nm,'DuPont_AT':dp_at,'DuPont_EM':dp_em,'DuPont_ROE':dp_roe,
        'CR':cr,'QR':qr,'CashR':cashr,'DE':de,'DTA':dta,'ND':nd,'ND_EB':nd_eb,'ICov':icov,
        'DSO':dso,'DIO':dio,'DPO':dpo,'CCC':ccc,'AT':at,'InvT':invt,'RecT':rect,
        'NWC':nwc,'NWC_P':nwc_p,
        'FixCost':fix,'VarCost':var,'Contrib':con,'ContribM':conm,'Breakeven':be,'OpLev':oplev,
        'FCFY':fcfy,'CFEB':cfeb,'CxR':cxr,'CxD':cxd,
        'NOPAT':nopat,'InvCap':ic,'EVA':eva,'Altman':altman,
        'RVar':rvar,'EVar':evar,'RMoM':rmom,'RYoY':ryoy,'EMoM':emom,'Seas':seas,
        'RevHC':rhc,'EBITDAHC':ehc,'OpExHC':ohc,
    })

df=load_data()

def mbb(fig,title='',h=370,yp='\u20ac',ys='',yf=','):
    fig.update_layout(
        title=dict(text=f'<b>{title}</b>',font=dict(size=13,color=NAVY,family='Inter'),x=0),
        font=dict(family='Inter',size=11,color=SLATE),plot_bgcolor='white',paper_bgcolor='white',
        margin=dict(l=48,r=16,t=48,b=36),
        legend=dict(orientation='h',yanchor='bottom',y=1.02,xanchor='right',x=1,font=dict(size=10),bgcolor='rgba(0,0,0,0)'),
        xaxis=dict(gridcolor='#F3F4F6',linecolor=BORDER,tickfont=dict(size=10)),
        yaxis=dict(gridcolor='#F3F4F6',linecolor=BORDER,tickprefix=yp,ticksuffix=ys,tickformat=yf,tickfont=dict(size=10)),
        hovermode='x unified',height=h,
        hoverlabel=dict(bgcolor='white',font_size=11,font_family='Inter',bordercolor=BORDER))
    return fig

def pcd(c,p): return f'{((c-p)/abs(p)*100):+.1f}%' if p!=0 else chr(8211)
def ppd(c,p): return f'{c-p:+.1f}pp'

with st.sidebar:
    st.markdown('#### \u25c8 GMA INNOVATION LAB')
    st.markdown('*Executive Financial Intelligence*')
    st.markdown('---')
    period=st.selectbox('REPORTING PERIOD',['Full Year 2024','H1 2024','H2 2024','Q1 2024','Q2 2024','Q3 2024','Q4 2024'])
    view=st.radio('ANALYSIS VIEW',[
        'Executive Summary','P&L Deep Dive','Revenue Intelligence',
        'Margin & Profitability','Balance Sheet','Cash Flow & FCF',
        'Working Capital','Ratios & Solvency','Variance & Growth'
    ])
    st.markdown('---')
    st.markdown(f'**Updated:** {datetime.now().strftime("%d %b %Y")}')
    st.markdown('**Source:** ERP + Accounting')
    st.markdown('**Currency:** EUR')
    st.markdown('---')
    st.caption('Sample data for demonstration.\nEach system customized to client.')

pm={'Full Year 2024':slice(None),'H1 2024':slice(0,6),'H2 2024':slice(6,12),'Q1 2024':slice(0,3),'Q2 2024':slice(3,6),'Q3 2024':slice(6,9),'Q4 2024':slice(9,12)}
d=df.iloc[pm[period]].reset_index(drop=True);n=len(d)
L=d.iloc[-1];P=d.iloc[-2] if n>1 else L

st.markdown('<p class="tag">GMA INNOVATION LAB \u2014 EXECUTIVE FINANCIAL SYSTEM</p>',unsafe_allow_html=True)
st.markdown(f'# Financial Intelligence \u2014 {period}')
st.markdown('<hr class="divider">',unsafe_allow_html=True)

if view=='Executive Summary':
    c1,c2,c3,c4,c5,c6=st.columns(6)
    c1.metric('Revenue',f'\u20ac{L["Revenue"]/1000:.0f}K',pcd(L['Revenue'],P['Revenue']))
    c2.metric('EBITDA Margin',f'{L["EM"]:.1f}%',ppd(L['EM'],P['EM']))
    c3.metric('Net Income',f'\u20ac{L["NI"]/1000:.0f}K',pcd(L['NI'],P['NI']))
    c4.metric('ROIC',f'{L["ROIC"]:.1f}%',ppd(L['ROIC'],P['ROIC']))
    c5.metric('FCF',f'\u20ac{L["FCF"]/1000:.0f}K',pcd(L['FCF'],P['FCF']))
    c6.metric('Altman Z',f'{L["Altman"]:.2f}','Safe' if L['Altman']>2.9 else 'Watch')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    rg=(d['Revenue'].iloc[-1]-d['Revenue'].iloc[0])/d['Revenue'].iloc[0]*100
    st.markdown(f'<div class="insight"><strong>Executive Insight:</strong> Revenue grew <strong>{rg:.1f}%</strong> over the period. Avg EBITDA margin: <strong>{d["EM"].mean():.1f}%</strong>. FCF: <strong>\u20ac{d["FCF"].sum()/1000:,.0f}K</strong> ({d["FCFY"].mean():.1f}% yield). ROIC <strong>{d["ROIC"].mean():.1f}%</strong> vs 10% WACC. CCC: <strong>{L["CCC"]:.0f}d</strong>. Altman Z: <strong>{L["Altman"]:.2f}</strong>.</div>',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Revenue']/1000,name='Revenue',marker_color=NAVY,opacity=0.85))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['EBITDA']/1000,name='EBITDA',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['NI']/1000,name='Net Income',line=dict(color=GREEN,width=2),mode='lines+markers',marker=dict(size=4)))
        fig=mbb(fig,'Revenue, EBITDA & Net Income (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        for col,nm2,clr in [('GM','Gross',NAVY),('EM','EBITDA',ACCENT),('NM','Net','#94A3B8')]:
            fig.add_trace(go.Scatter(x=d['Label'],y=d[col],name=nm2,line=dict(color=clr,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Margin Cascade (%)');fig.update_yaxes(tickprefix='',ticksuffix='%',tickformat='.1f')
        st.plotly_chart(fig,use_container_width=True)
    c3,c4=st.columns(2)
    with c3:
        fig=go.Figure(go.Waterfall(x=['Revenue','COGS','GP','OpEx','EBITDA','D&A','Int/Tax','Net Inc'],y=[L['Revenue'],-L['COGS'],0,-L['TotalOpEx'],0,-L['DA'],-L['Interest']-L['Tax'],0],measure=['absolute','relative','total','relative','total','relative','relative','total'],connector=dict(line=dict(color=BORDER,width=1)),increasing=dict(marker=dict(color=NAVY)),decreasing=dict(marker=dict(color=RED)),totals=dict(marker=dict(color=ACCENT)),textposition='outside',text=[f'\u20ac{v/1000:.0f}K' for v in [L['Revenue'],L['COGS'],L['GP'],L['TotalOpEx'],L['EBITDA'],L['DA'],L['Interest']+L['Tax'],L['NI']]],textfont=dict(size=9,color=SLATE)))
        fig=mbb(fig,f'P&L Bridge \u2014 {L["Label"]}',h=380);fig.update_yaxes(visible=False)
        st.plotly_chart(fig,use_container_width=True)
    with c4:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ROIC'],name='ROIC',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ROE'],name='ROE',line=dict(color=ACCENT,width=2),mode='lines+markers',marker=dict(size=4)))
        fig.add_hline(y=10,line=dict(color=RED,dash='dash',width=1),annotation_text='WACC 10%')
        fig=mbb(fig,'Return Metrics vs WACC (%)');fig.update_yaxes(tickprefix='',ticksuffix='%',tickformat='.1f')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### Executive KPI Dashboard')
    kpi_data=[]
    for name,val,tgt,good in [('Gross Margin',f'{L["GM"]:.1f}%','> 40%',L['GM']>40),('EBITDA Margin',f'{L["EM"]:.1f}%','> 18%',L['EM']>18),('ROIC',f'{L["ROIC"]:.1f}%','> WACC',L['ROIC']>10),('Current Ratio',f'{L["CR"]:.2f}x','> 1.5x',L['CR']>1.5),('CCC',f'{L["CCC"]:.0f}d','< 45d',L['CCC']<45),('Interest Cov',f'{L["ICov"]:.1f}x','> 3x',L['ICov']>3),('FCF Yield',f'{L["FCFY"]:.1f}%','> 5%',L['FCFY']>5),('Altman Z',f'{L["Altman"]:.2f}','> 2.9',L['Altman']>2.9)]:
        kpi_data.append({'KPI':name,'Current':val,'Target':tgt,'Status':'\u2705 On Target' if good else '\u26a0\ufe0f Monitor'})
    st.dataframe(pd.DataFrame(kpi_data),use_container_width=True,hide_index=True)

elif view=='P&L Deep Dive':
    c1,c2,c3,c4=st.columns(4)
    c1.metric('Revenue',f'\u20ac{d["Revenue"].sum()/1000:,.0f}K')
    c2.metric('Gross Profit',f'\u20ac{d["GP"].sum()/1000:,.0f}K',f'{d["GM"].mean():.1f}% margin')
    c3.metric('EBITDA',f'\u20ac{d["EBITDA"].sum()/1000:,.0f}K',f'{d["EM"].mean():.1f}% margin')
    c4.metric('Net Income',f'\u20ac{d["NI"].sum()/1000:,.0f}K',f'{d["NM"].mean():.1f}% margin')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Revenue']/1000,name='Revenue',marker_color=NAVY,opacity=0.85))
        fig.add_trace(go.Bar(x=d['Label'],y=-d['COGS']/1000,name='COGS',marker_color=RED,opacity=0.7))
        fig.add_trace(go.Bar(x=d['Label'],y=-d['TotalOpEx']/1000,name='OpEx',marker_color=AMBER,opacity=0.7))
        fig=mbb(fig,'Revenue vs Cost Structure (\u20acK)');fig.update_layout(barmode='relative');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        for col,nm2,clr in [('Pers','Personnel',NAVY),('Rent','Rent',ACCENT),('Mkt','Marketing','#94A3B8'),('Trav','Travel',TEAL),('Tech','Technology',PURPLE),('Prof','Professional',AMBER),('OOx','Other','#CBD5E1')]:
            fig.add_trace(go.Bar(x=d['Label'],y=d[col]/1000,name=nm2,marker_color=clr))
        fig=mbb(fig,'OpEx Breakdown (\u20acK)');fig.update_layout(barmode='stack');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### P&L Statement')
    pl=pd.DataFrame({'Month':d['Label'],'Revenue':d['Revenue'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'COGS':d['COGS'].apply(lambda x:f'(\u20ac{x/1000:,.0f}K)'),'GP':d['GP'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'GM%':d['GM'].apply(lambda x:f'{x:.1f}%'),'OpEx':d['TotalOpEx'].apply(lambda x:f'(\u20ac{x/1000:,.0f}K)'),'EBITDA':d['EBITDA'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'EM%':d['EM'].apply(lambda x:f'{x:.1f}%'),'EBIT':d['EBIT'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'NI':d['NI'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'NM%':d['NM'].apply(lambda x:f'{x:.1f}%')})
    st.dataframe(pl,use_container_width=True,hide_index=True)

elif view=='Revenue Intelligence':
    c1,c2,c3,c4=st.columns(4)
    c1.metric('Product Rev',f'\u20ac{d["Rev_Prod"].sum()/1000:,.0f}K',f'{d["Rev_Prod"].sum()/d["Revenue"].sum()*100:.0f}% mix')
    c2.metric('Service Rev',f'\u20ac{d["Rev_Svc"].sum()/1000:,.0f}K',f'{d["Rev_Svc"].sum()/d["Revenue"].sum()*100:.0f}% mix')
    c3.metric('Rev/Employee',f'\u20ac{L["RevHC"]/1000:.0f}K',pcd(L['RevHC'],P['RevHC']))
    c4.metric('Seasonality',f'{L["Seas"]:.2f}x','Above avg' if L['Seas']>1 else 'Below avg')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Rev_Prod']/1000,name='Product',marker_color=NAVY))
        fig.add_trace(go.Bar(x=d['Label'],y=d['Rev_Svc']/1000,name='Service',marker_color=ACCENT))
        fig.add_trace(go.Bar(x=d['Label'],y=d['Rev_Oth']/1000,name='Other',marker_color='#94A3B8'))
        fig=mbb(fig,'Revenue by Segment (\u20acK)');fig.update_layout(barmode='stack');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Seas'],name='Seasonality',marker_color=[NAVY if s>=1 else '#94A3B8' for s in d['Seas']]))
        fig.add_hline(y=1.0,line=dict(color=RED,dash='dash',width=1),annotation_text='Avg = 1.0x')
        fig=mbb(fig,'Revenue Seasonality Index');fig.update_yaxes(tickprefix='',ticksuffix='x',tickformat='.2f')
        st.plotly_chart(fig,use_container_width=True)
    c3,c4=st.columns(2)
    with c3:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['RevHC']/1000,name='Rev/HC',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['EBITDAHC']/1000,name='EBITDA/HC',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Productivity per Employee (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c4:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['RYoY'],name='YoY',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Bar(x=d['Label'],y=d['RMoM'],name='MoM',marker_color=ACCENT,opacity=0.5))
        fig=mbb(fig,'Revenue Growth (%)');fig.update_yaxes(tickprefix='',ticksuffix='%')
        st.plotly_chart(fig,use_container_width=True)

elif view=='Margin & Profitability':
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric('Gross',f'{L["GM"]:.1f}%',f'\u03c3={d["GM"].std():.1f}pp')
    c2.metric('EBITDA',f'{L["EM"]:.1f}%',f'\u03c3={d["EM"].std():.1f}pp')
    c3.metric('Net',f'{L["NM"]:.1f}%',f'\u03c3={d["NM"].std():.1f}pp')
    c4.metric('ROIC',f'{L["ROIC"]:.1f}%',ppd(L['ROIC'],P['ROIC']))
    c5.metric('EVA',f'\u20ac{L["EVA"]/1000:.0f}K','Value +' if L['EVA']>0 else 'Value \u2212')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=make_subplots(specs=[[{'secondary_y':True}]])
        fig.add_trace(go.Bar(x=d['Label'],y=d['Revenue']/1000,name='Revenue',marker_color=NAVY,opacity=0.25),secondary_y=False)
        fig.add_trace(go.Scatter(x=d['Label'],y=d['GM'],name='Gross',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)),secondary_y=True)
        fig.add_trace(go.Scatter(x=d['Label'],y=d['EM'],name='EBITDA',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)),secondary_y=True)
        fig.add_trace(go.Scatter(x=d['Label'],y=d['NM'],name='Net',line=dict(color='#94A3B8',width=2),mode='lines+markers',marker=dict(size=4)),secondary_y=True)
        fig=mbb(fig,'Scale vs Margins');fig.update_yaxes(tickprefix='\u20ac',ticksuffix='K',secondary_y=False);fig.update_yaxes(ticksuffix='%',tickformat='.1f',secondary_y=True,gridcolor='rgba(0,0,0,0)')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ROIC'],name='ROIC',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ROE'],name='ROE',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ROA'],name='ROA',line=dict(color=TEAL,width=2),mode='lines+markers',marker=dict(size=4)))
        fig.add_hline(y=10,line=dict(color=RED,dash='dash',width=1),annotation_text='WACC 10%')
        fig=mbb(fig,'Return Metrics (%)');fig.update_yaxes(tickprefix='',ticksuffix='%',tickformat='.1f')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### DuPont Decomposition \u2014 ROE = Margin \u00d7 Turnover \u00d7 Leverage')
    c1,c2,c3=st.columns(3)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['DuPont_NM']*100,name='Net Margin',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Net Margin (Profitability)');fig.update_yaxes(tickprefix='',ticksuffix='%')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['DuPont_AT'],name='Asset Turnover',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Asset Turnover (Efficiency)');fig.update_yaxes(tickprefix='',ticksuffix='x',tickformat='.2f')
        st.plotly_chart(fig,use_container_width=True)
    with c3:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['DuPont_EM'],name='Equity Multiplier',line=dict(color=TEAL,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Equity Multiplier (Leverage)');fig.update_yaxes(tickprefix='',ticksuffix='x',tickformat='.2f')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### Operating Leverage & Breakeven')
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['FixCost']/1000,name='Fixed',marker_color=NAVY))
        fig.add_trace(go.Bar(x=d['Label'],y=d['VarCost']/1000,name='Variable',marker_color=ACCENT))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['Revenue']/1000,name='Revenue',line=dict(color=GREEN,width=2,dash='dot'),mode='lines'))
        fig=mbb(fig,'Cost Structure vs Revenue (\u20acK)');fig.update_layout(barmode='stack');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Revenue']/1000,name='Revenue',marker_color=NAVY,opacity=0.3))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['Breakeven']/1000,name='Breakeven',line=dict(color=RED,width=2.5,dash='dash'),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Revenue vs Breakeven (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)

elif view=='Balance Sheet':
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric('Total Assets',f'\u20ac{L["TA"]/1000:,.0f}K',pcd(L['TA'],P['TA']))
    c2.metric('Current Ratio',f'{L["CR"]:.2f}x',f'{L["CR"]-P["CR"]:+.2f}')
    c3.metric('Quick Ratio',f'{L["QR"]:.2f}x',f'{L["QR"]-P["QR"]:+.2f}')
    c4.metric('Debt/Equity',f'{L["DE"]:.2f}x',f'{L["DE"]-P["DE"]:+.2f}',delta_color='inverse')
    c5.metric('ND/EBITDA',f'{L["ND_EB"]:.1f}x',f'{L["ND_EB"]-P["ND_EB"]:+.1f}',delta_color='inverse')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Cash']/1000,name='Cash',marker_color=GREEN))
        fig.add_trace(go.Bar(x=d['Label'],y=d['Rec_T']/1000,name='Receivables',marker_color=NAVY))
        fig.add_trace(go.Bar(x=d['Label'],y=d['Inv']/1000,name='Inventory',marker_color=ACCENT))
        fig=mbb(fig,'Current Asset Composition (\u20acK)');fig.update_layout(barmode='stack');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['CR'],name='Current',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['QR'],name='Quick',line=dict(color=ACCENT,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['CashR'],name='Cash',line=dict(color=TEAL,width=2),mode='lines+markers',marker=dict(size=4)))
        fig.add_hline(y=1.5,line=dict(color=GREEN,dash='dash',width=1),annotation_text='Target 1.5x')
        fig=mbb(fig,'Liquidity Ratios');fig.update_yaxes(tickprefix='',ticksuffix='x',tickformat='.2f')
        st.plotly_chart(fig,use_container_width=True)

elif view=='Cash Flow & FCF':
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric('Operating CF',f'\u20ac{L["OpCF"]/1000:.0f}K',pcd(L['OpCF'],P['OpCF']))
    c2.metric('FCF',f'\u20ac{L["FCF"]/1000:.0f}K',pcd(L['FCF'],P['FCF']))
    c3.metric('FCF Yield',f'{L["FCFY"]:.1f}%',ppd(L['FCFY'],P['FCFY']))
    c4.metric('CF/EBITDA',f'{L["CFEB"]:.0f}%',ppd(L['CFEB'],P['CFEB']))
    c5.metric('CapEx/Rev',f'{L["CxR"]:.1f}%')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['OpCF']/1000,name='Operating',marker_color=NAVY))
        fig.add_trace(go.Bar(x=d['Label'],y=d['InvCF']/1000,name='Investing',marker_color=RED,opacity=0.7))
        fig.add_trace(go.Bar(x=d['Label'],y=d['FinCF']/1000,name='Financing',marker_color='#94A3B8'))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['NetCF']/1000,name='Net',line=dict(color=ACCENT,width=2.5,dash='dot'),mode='lines+markers',marker=dict(size=6)))
        fig=mbb(fig,'Cash Flow Components (\u20acK)');fig.update_layout(barmode='group');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['OpCF']/1000,name='Op CF',marker_color=NAVY,opacity=0.3))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['FCF']/1000,name='FCF',fill='tozeroy',fillcolor='rgba(5,150,105,0.08)',line=dict(color=GREEN,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Free Cash Flow (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown(f'<div class="insight"><strong>Cash Quality:</strong> OpCF/EBITDA = <strong>{d["CFEB"].mean():.0f}%</strong> avg. CapEx intensity: <strong>{d["CxR"].mean():.1f}%</strong> of revenue. CapEx/Dep: <strong>{d["CxD"].mean():.1f}x</strong>. Cumulative FCF: <strong>\u20ac{d["FCF"].sum()/1000:,.0f}K</strong>.</div>',unsafe_allow_html=True)

elif view=='Working Capital':
    c1,c2,c3,c4,c5=st.columns(5)
    c1.metric('DSO',f'{L["DSO"]:.0f}d',f'{L["DSO"]-P["DSO"]:+.0f}d',delta_color='inverse')
    c2.metric('DIO',f'{L["DIO"]:.0f}d',f'{L["DIO"]-P["DIO"]:+.0f}d',delta_color='inverse')
    c3.metric('DPO',f'{L["DPO"]:.0f}d',f'{L["DPO"]-P["DPO"]:+.0f}d')
    c4.metric('CCC',f'{L["CCC"]:.0f}d',f'{L["CCC"]-P["CCC"]:+.0f}d',delta_color='inverse')
    c5.metric('NWC % Rev',f'{L["NWC_P"]:.1f}%',ppd(L['NWC_P'],P['NWC_P']),delta_color='inverse')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        for col,nm2,clr in [('DSO','DSO',NAVY),('DIO','DIO',ACCENT),('DPO','DPO',GREEN)]:
            fig.add_trace(go.Scatter(x=d['Label'],y=d[col],name=nm2,line=dict(color=clr,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'Working Capital Cycle (Days)');fig.update_yaxes(tickprefix='',ticksuffix='d')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['CCC'],fill='tozeroy',fillcolor='rgba(27,42,74,0.06)',line=dict(color=NAVY,width=2.5),name='CCC',mode='lines+markers',marker=dict(size=5)))
        fig.add_hline(y=d['CCC'].mean(),line=dict(color=ACCENT,dash='dash',width=1),annotation_text=f'Avg {d["CCC"].mean():.0f}d')
        fig=mbb(fig,'Cash Conversion Cycle (Days)');fig.update_yaxes(tickprefix='',ticksuffix='d')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### Net Working Capital Bridge')
    fig=go.Figure()
    fig.add_trace(go.Bar(x=d['Label'],y=d['Rec_T']/1000,name='Receivables',marker_color=NAVY))
    fig.add_trace(go.Bar(x=d['Label'],y=d['Inv']/1000,name='Inventory',marker_color=ACCENT))
    fig.add_trace(go.Bar(x=d['Label'],y=-d['Pay_T']/1000,name='Payables',marker_color=GREEN))
    fig.add_trace(go.Scatter(x=d['Label'],y=d['NWC']/1000,name='Net WC',line=dict(color=RED,width=2.5,dash='dot'),mode='lines+markers',marker=dict(size=6)))
    fig=mbb(fig,'Working Capital Components (\u20acK)',h=340);fig.update_layout(barmode='relative');fig.update_yaxes(ticksuffix='K')
    st.plotly_chart(fig,use_container_width=True)

elif view=='Ratios & Solvency':
    c1,c2,c3,c4=st.columns(4)
    c1.metric('Altman Z-Score',f'{L["Altman"]:.2f}','Safe' if L['Altman']>2.9 else 'Grey Zone')
    c2.metric('Interest Coverage',f'{L["ICov"]:.1f}x',f'{L["ICov"]-P["ICov"]:+.1f}')
    c3.metric('Net Debt',f'\u20ac{L["ND"]/1000:.0f}K')
    c4.metric('Debt/Assets',f'{L["DTA"]:.1%}',f'{(L["DTA"]-P["DTA"])*100:+.1f}pp',delta_color='inverse')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['Altman'],name='Z-Score',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=6)))
        fig.add_hline(y=2.9,line=dict(color=GREEN,dash='dash',width=1),annotation_text='Safe > 2.9')
        fig.add_hline(y=1.8,line=dict(color=RED,dash='dash',width=1),annotation_text='Distress < 1.8')
        fig.add_hrect(y0=1.8,y1=2.9,fillcolor=AMBER,opacity=0.05)
        fig=mbb(fig,'Altman Z-Score (Private)');fig.update_yaxes(tickprefix='')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=d['Label'],y=d['ICov'],name='Interest Coverage',line=dict(color=NAVY,width=2.5),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['DE'],name='Debt/Equity',line=dict(color=ACCENT,width=2),mode='lines+markers',marker=dict(size=4)))
        fig.add_hline(y=3.0,line=dict(color=GREEN,dash='dash',width=1),annotation_text='Coverage 3x')
        fig=mbb(fig,'Leverage & Coverage');fig.update_yaxes(tickprefix='',ticksuffix='x')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### Economic Value Added (EVA)')
    fig=go.Figure()
    colors=[GREEN if v>0 else RED for v in d['EVA']]
    fig.add_trace(go.Bar(x=d['Label'],y=d['EVA']/1000,name='EVA',marker_color=colors))
    fig=mbb(fig,'Monthly EVA \u2014 Value Creation vs Destruction (\u20acK)');fig.update_yaxes(ticksuffix='K')
    st.plotly_chart(fig,use_container_width=True)
    st.markdown(f'<div class="insight"><strong>EVA:</strong> NOPAT avg <strong>\u20ac{d["NOPAT"].mean()/1000:,.0f}K</strong>/mo vs capital charge <strong>\u20ac{(0.10/12*d["InvCap"].mean())/1000:,.0f}K</strong>/mo. Cumulative EVA: <strong>\u20ac{d["EVA"].sum()/1000:,.0f}K</strong>.</div>',unsafe_allow_html=True)

elif view=='Variance & Growth':
    rv=d['RVar'].mean();ev=d['EVar'].mean()
    c1,c2,c3,c4=st.columns(4)
    c1.metric('Rev vs Budget',f'{rv:+.1f}%','Above' if rv>0 else 'Below')
    c2.metric('EBITDA vs Budget',f'{ev:+.1f}%','Above' if ev>0 else 'Below')
    c3.metric('Rev YoY',f'{L["RYoY"]:.1f}%')
    c4.metric('Best Month',f'{d.loc[d["Revenue"].idxmax(),"Label"]}')
    st.markdown('<hr class="divider">',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['Revenue']/1000,name='Actual',marker_color=NAVY,opacity=0.85))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['BRev']/1000,name='Budget',line=dict(color=RED,width=2,dash='dash'),mode='lines+markers',marker=dict(size=5)))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['PYRev']/1000,name='Prior Year',line=dict(color='#94A3B8',width=1.5,dash='dot'),mode='lines'))
        fig=mbb(fig,'Revenue: Actual vs Budget vs PY (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    with c2:
        fig=go.Figure()
        fig.add_trace(go.Bar(x=d['Label'],y=d['EBITDA']/1000,name='Actual',marker_color=NAVY,opacity=0.85))
        fig.add_trace(go.Scatter(x=d['Label'],y=d['BEBITDA']/1000,name='Budget',line=dict(color=RED,width=2,dash='dash'),mode='lines+markers',marker=dict(size=5)))
        fig=mbb(fig,'EBITDA: Actual vs Budget (\u20acK)');fig.update_yaxes(ticksuffix='K')
        st.plotly_chart(fig,use_container_width=True)
    st.markdown('### Variance Detail')
    vdf=pd.DataFrame({'Month':d['Label'],'Rev':d['Revenue'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'Budget':d['BRev'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'\u0394%':d['RVar'].apply(lambda x:f'{x:+.1f}%'),'PY':d['PYRev'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'YoY':d['RYoY'].apply(lambda x:f'{x:+.1f}%'),'EBITDA':d['EBITDA'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'Bgt':d['BEBITDA'].apply(lambda x:f'\u20ac{x/1000:,.0f}K'),'E\u0394%':d['EVar'].apply(lambda x:f'{x:+.1f}%')})
    st.dataframe(vdf,use_container_width=True,hide_index=True)

st.markdown('<hr class="divider">',unsafe_allow_html=True)

def make_excel(data):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        cols_pl=['Label','Revenue','COGS','GP','TotalOpEx','EBITDA','DA','EBIT','Interest','EBT','Tax','NI','GM','EM','EBITM','NM']
        data[cols_pl].rename(columns={'Label':'Month','GP':'Gross Profit','TotalOpEx':'OpEx','DA':'D&A','GM':'GM%','EM':'EM%','EBITM':'EBIT%','NM':'NM%'}).to_excel(w,sheet_name='P&L',index=False)
        cols_bs=['Label','Cash','Rec_T','Inv','TCA','Pay','SD','TCL','LD','TNCL','TL','Eq','TA']
        data[cols_bs].rename(columns={'Label':'Month','Rec_T':'Receivables','Pay':'Payables','SD':'Short Debt','LD':'Long Debt','Eq':'Equity','TA':'Total Assets','TCA':'Current Assets','TCL':'Current Liab','TNCL':'Non-Current Liab','TL':'Total Liab'}).to_excel(w,sheet_name='Balance Sheet',index=False)
        cols_cf=['Label','OpCF','CapEx','InvCF','FinCF','NetCF','FCF','Cash']
        data[cols_cf].rename(columns={'Label':'Month','OpCF':'Operating CF','InvCF':'Investing CF','FinCF':'Financing CF','NetCF':'Net CF','FCF':'Free CF'}).to_excel(w,sheet_name='Cash Flow',index=False)
        cols_r=['Label','GM','EM','NM','ROE','ROA','ROIC','CR','QR','DE','ICov','DSO','DIO','DPO','CCC','FCFY','CFEB','Altman','NWC_P','ContribM','OpLev']
        data[cols_r].rename(columns={'Label':'Month','GM':'Gross%','EM':'EBITDA%','NM':'Net%','CR':'Current','QR':'Quick','DE':'D/E','ICov':'IntCov','FCFY':'FCF Yield','CFEB':'CF/EBITDA','NWC_P':'NWC/Rev','ContribM':'Contrib%','OpLev':'OpLev'}).to_excel(w,sheet_name='Ratios',index=False)
        for sn in w.sheets:
            ws2=w.sheets[sn]
            for cell in ws2[1]:
                cell.font=Font(name='Calibri',size=10,bold=True,color='FFFFFF')
                cell.fill=PatternFill(start_color='1B2A4A',end_color='1B2A4A',fill_type='solid')
                cell.alignment=Alignment(horizontal='center')
            for col in ws2.columns: ws2.column_dimensions[get_column_letter(col[0].column)].width=15
    buf.seek(0);return buf

c1,c2,_=st.columns([1,1,4])
with c1:
    st.download_button('\U0001f4e5 Export Report',make_excel(d),file_name=f'gma_report_{period.replace(" ","_").lower()}.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',use_container_width=True)
with c2:
    tpl=os.path.join(os.path.dirname(os.path.abspath(__file__)),'gma_financial_template.xlsx')
    if os.path.exists(tpl):
        with open(tpl,'rb') as f:
            st.download_button('\U0001f4cb Client Template',f.read(),file_name='gma_financial_template.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',use_container_width=True)

st.markdown('---')
st.caption('\u25c8 GMA Innovation Lab \u2014 Executive Financial Intelligence System | MBB Methodology')
